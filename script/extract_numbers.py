import re
from openpyxl import load_workbook


def extract_numbers_from_excel(file_path, column_letter):
    # 加载工作簿
    wb = load_workbook(file_path)

    # 选择活动工作表
    ws = wb.active

    # 获取指定列的所有单元格
    column = ws[column_letter]

    # 遍历列中的每个单元格
    for cell in column:
        if cell.value:
            # 使用正则表达式提取数字
            numbers = re.sub(r"[^\d,.-]", "", cell.value.replace(",", "."))
            # 如果找到数字，则将它们连接起来并写回单元格
            if numbers:
                numbers = round(float(numbers) * 0.0096, 2)
                cell.value = numbers

    # 保存修改后的工作簿
    wb.save(file_path)
    print(f"数字已提取并写回 {file_path} 的 {column_letter} 列。")


# 使用示例
file_path = "/Users/changtong/Downloads/all-en(for import).xlsx"
column_letter = "E"  # 替换为您想要处理的列字母

extract_numbers_from_excel(file_path, column_letter)
